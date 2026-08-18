"""XLSX: hojas de calculo (soporte basico y honesto).

Se procesan valores de texto; los numeros se copian TAL CUAL para no tocar
resultados. No se copian formulas, macros, graficos, imagenes ni formato.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from ..models import Alerta, Capa, DocumentoExtraido
from .base import ManejadorFormato, hallazgos_de_metadatos, unidad

CAMPOS_PROPS = [
    "creator", "title", "description", "subject", "identifier", "language",
    "keywords", "category", "contentStatus", "version", "revision",
    "lastModifiedBy", "company", "manager",
]


class ManejadorXlsx(ManejadorFormato):
    extensiones = (".xlsx",)
    nombre = "xlsx"
    reconstruccion = "parcial"
    capas_no_copiadas = frozenset({Capa.COMENTARIO})

    def extraer(self, ruta: Path) -> DocumentoExtraido:
        ruta = Path(ruta)
        doc = DocumentoExtraido(formato="xlsx")
        try:
            wb = load_workbook(str(ruta), data_only=False, keep_vba=False)
        except Exception as exc:
            doc.alertas.append(
                Alerta("critica", "XLSX_ILEGIBLE", "No se pudo abrir el libro",
                       str(exc)[:200])
            )
            return doc

        metadatos = {}
        props = wb.properties
        for campo in CAMPOS_PROPS:
            valor = getattr(props, campo, None)
            if valor:
                metadatos["props:" + campo] = str(valor)[:200]
        try:
            for prop in (wb.custom_doc_props or []):
                if getattr(prop, "value", None):
                    metadatos["custom:" + str(prop.name)] = str(prop.value)[:200]
        except Exception:
            pass

        estructura = []
        formulas = 0
        comentarios = 0
        for s, hoja in enumerate(wb.worksheets):
            doc.unidades.append(
                unidad("hoja%d" % s, hoja.title, Capa.HOJA_CALCULO,
                       "nombre de hoja[%d]" % s)
            )
            celdas = []
            for fila in hoja.iter_rows():
                for celda in fila:
                    if celda.value is None:
                        continue
                    if celda.comment is not None:
                        comentarios += 1
                        doc.unidades.append(
                            unidad("com%d_%s" % (s, celda.coordinate),
                                   str(celda.comment.text), Capa.COMENTARIO,
                                   "hoja[%d].%s.comentario" % (s, celda.coordinate),
                                   editable=False)
                        )
                    if isinstance(celda.value, str):
                        if celda.value.startswith("="):
                            formulas += 1
                            continue
                        uid = "c%d_%s" % (s, celda.coordinate)
                        doc.unidades.append(
                            unidad(uid, celda.value, Capa.HOJA_CALCULO,
                                   "hoja[%d].%s" % (s, celda.coordinate))
                        )
                        celdas.append({"uid": uid, "coord": celda.coordinate,
                                       "tipo": "texto"})
                    else:
                        # Los numeros NO se transforman, pero si se registran
                        # como unidad para poder verificar que sobreviven
                        # intactos a la reconstruccion.
                        uid = "v%d_%s" % (s, celda.coordinate)
                        doc.unidades.append(
                            unidad(uid, str(celda.value), Capa.HOJA_CALCULO,
                                   "hoja[%d].%s" % (s, celda.coordinate),
                                   editable=False)
                        )
                        celdas.append({"coord": celda.coordinate, "tipo": "valor",
                                       "valor": celda.value})
            estructura.append({"hoja_uid": "hoja%d" % s, "titulo": hoja.title,
                               "celdas": celdas})
        wb.close()

        doc.metadatos = metadatos
        doc.hallazgos_tecnicos = hallazgos_de_metadatos(metadatos)
        doc.unidades.append(
            unidad("nombre_archivo", ruta.stem, Capa.NOMBRE_ARCHIVO,
                   "nombre del archivo")
        )
        doc.info = {"hojas": len(estructura), "estructura": estructura,
                    "formulas": formulas, "comentarios": comentarios}
        if formulas:
            doc.alertas.append(
                Alerta("advertencia", "FORMULAS",
                       "El libro tiene %d formula(s)." % formulas,
                       "No se copian al resultado: podrian referirse a datos "
                       "que ya no existen.")
            )
        return doc

    def reconstruir(self, extraido, unidades_nuevas, destino: Path, contexto=None):
        mapa = {u.uid: u for u in unidades_nuevas}
        wb = Workbook()
        wb.remove(wb.active)
        esperado = []
        for hoja_info in extraido.info.get("estructura", []):
            titulo_u = mapa.get(hoja_info["hoja_uid"])
            titulo = (titulo_u.texto if titulo_u else hoja_info["titulo"]) or "Hoja"
            titulo = titulo[:31].replace("/", "-").replace("\\", "-").replace("*", "")
            titulo = titulo.replace("[", "(").replace("]", ")").replace("?", "")
            hoja = wb.create_sheet(title=titulo or "Hoja")
            esperado.append(titulo)
            for celda in hoja_info["celdas"]:
                if celda["tipo"] == "texto":
                    u = mapa.get(celda["uid"])
                    valor = u.texto if u else ""
                    hoja[celda["coord"]] = valor
                    esperado.append(valor)
                else:
                    hoja[celda["coord"]] = celda["valor"]
                    esperado.append(str(celda["valor"]))
        if not wb.worksheets:
            wb.create_sheet(title="Hoja")

        props = wb.properties
        props.creator = ""
        props.lastModifiedBy = ""
        props.title = "Libro desidentificado"
        props.description = "Generado por ANONIMIZADOR. Requiere revision humana."
        props.subject = ""
        props.keywords = ""
        props.category = ""
        props.identifier = ""
        props.language = ""
        props.contentStatus = "DESIDENTIFICADO - PENDIENTE DE REVISION"
        props.revision = None
        neutro = datetime(2000, 1, 1, 0, 0, 0)
        props.created = neutro
        props.modified = neutro
        wb.save(str(destino))
        return {
            "texto_esperado": "\n".join(esperado),
            "capas_descartadas": ["formulas", "comentarios de celda", "formato",
                                  "graficos e imagenes"],
            "notas": [
                "Libro reconstruido con openpyxl: solo valores.",
                "Los numeros se copian sin tocarlos.",
            ],
        }
